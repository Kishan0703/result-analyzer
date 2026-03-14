# backend/services/excel_generator.py
"""Generate section-wise Excel workbook with one sheet per section plus combined."""

from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.analyzer import build_full_report
from services.merger import detect_grouping


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
SECTION_FONT = Font(bold=True, size=11, color="1E3A5F")
ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_SIDE = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _apply_data_style(cell, alternate=False, left=False):
    if alternate:
        cell.fill = ALT_FILL
    cell.alignment = LEFT if left else CENTER
    cell.border = BORDER


def _normalize_category_for_excel(value: object) -> str:
    key = str(value or "").strip().upper()
    mapping = {
        "CET": "CET+SNQ",
        "SNQ": "CET+SNQ",
        "COMED-K": "Comed-K",
        "COMEDK": "Comed-K",
        "MGMT": "MNG+PIO+JK",
        "MANAGEMENT": "MNG+PIO+JK",
        "MNG+PIO+JK": "MNG+PIO+JK",
        "PIO": "MNG+PIO+JK",
        "DIP": "DIP",
        "DIPLOMA": "DIP",
    }
    return mapping.get(key, "")


def _build_category_rows(category_wise: list[dict] | None) -> list[list]:
    rows = [["CET+SNQ", 0, 0, 0], ["Comed-K", 0, 0, 0], ["MNG+PIO+JK", 0, 0, 0], ["DIP", 0, 0, 0]]
    if not category_wise:
        return rows

    idx = {"CET+SNQ": 0, "Comed-K": 1, "MNG+PIO+JK": 2, "DIP": 3}
    for item in category_wise:
        label = _normalize_category_for_excel(item.get("category"))
        if not label:
            continue
        pos = idx[label]
        rows[pos][1] += int(item.get("pass", 0) or 0)
        rows[pos][2] += int(item.get("fail", 0) or 0)
        rows[pos][3] += int(item.get("total", 0) or 0)

    return rows


def write_section_to_sheet(ws, report, group_label):
    """Write all four report parts into one worksheet, stacked vertically."""
    current_row = 1

    def write_title(text):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=text).font = SECTION_FONT
        current_row += 1

    def write_header(headers):
        nonlocal current_row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        current_row += 1

    def write_row(values, alternate=False, left_cols=None):
        nonlocal current_row
        left_cols = left_cols or set()
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=v)
            _apply_data_style(cell, alternate=alternate, left=(col in left_cols))
        current_row += 1

    def blank_row():
        nonlocal current_row
        current_row += 1

    ws.cell(row=current_row, column=1, value=f"Result Analysis - {group_label}").font = Font(bold=True, size=13)
    current_row += 2

    write_title("1. Subject-wise Result Analysis")
    headers_1 = [
        "Sl No",
        "Subject Title",
        "Subject Code",
        "Staff Name",
        "Class Strength",
        "Appeared",
        "Passed",
        "AB",
        "DX",
        "NP",
        "Failed",
        "With Held",
        "Pass %",
    ]
    write_header(headers_1)
    for i, s in enumerate(report["subjects"]):
        write_row(
            [
                s.get("sl_no"),
                s.get("course_title"),
                s.get("course_code"),
                s.get("staff_name"),
                s.get("class_strength"),
                s.get("appeared"),
                s.get("passed"),
                s.get("absent"),
                s.get("dx"),
                s.get("np"),
                s.get("failed"),
                "",
                f"{s.get('pass_percentage', 0)}%",
            ],
            alternate=(i % 2 == 1),
            left_cols={2, 4},
        )
    blank_row()

    write_title("2. Topper's List")
    headers_2 = ["Sl.No", "USN", "Name", "SGPA", "CGPA"]
    write_header(headers_2)
    for i, t in enumerate(report["toppers"]):
        write_row(
            [t.get("rank"), t.get("usn"), t.get("name"), t.get("sgpa"), t.get("cgpa")],
            alternate=(i % 2 == 1),
            left_cols={3},
        )
    blank_row()

    write_title("3. Overall Category-wise Result")
    headers_3 = ["Quota", "Pass", "Fail", "Total"]
    write_header(headers_3)
    for i, row in enumerate(_build_category_rows(report.get("category_wise"))):
        write_row(row, alternate=(i % 2 == 1), left_cols={1})
    blank_row()

    write_title("4. Backlog Analysis")
    headers_4 = ["Backlog Category", "No. of Students"]
    write_header(headers_4)
    dist = report["backlog"]["distribution"]
    labels = [
        ("Single Backlog", "1"),
        ("Two Backlog", "2"),
        ("Three Backlog", "3"),
        ("Four Backlog", "4"),
        ("Five Backlog", "5"),
        ("Six & above Backlog", "6+"),
    ]
    for i, (label, key) in enumerate(labels):
        write_row([label, dist.get(key, 0)], alternate=(i % 2 == 1), left_cols={1})

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)


def generate_excel(df, enrichment, metadata, output_path, group_filter=None):
    """Generate one workbook with one sheet per section and a combined sheet."""
    df = detect_grouping(df)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if group_filter:
        groups = [group_filter]
    else:
        groups = sorted(df["group"].dropna().astype(str).unique().tolist())
        groups.append("ALL")

    for group in groups:
        if group == "ALL":
            group_df = df
            sheet_name = "Combined"
            label = "All Sections"
        else:
            group_df = df[df["group"].astype(str) == str(group)]
            sheet_name = f"{group}-Sec"
            label = f"{group}-Sec"

        report = build_full_report(group_df, enrichment)
        ws = wb.create_sheet(sheet_name[:31])
        write_section_to_sheet(ws, report, label)

    wb.save(output_path)
